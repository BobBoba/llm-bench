#!/usr/bin/env python3
# Appends LOCAL LM Studio model rows to "AI price comparison.xlsx" (RUST + humanitarian),
# mirroring the column layout of the cloud ZDR benchmark.
#
#   * Merges ALL results/*rust*.json and results/*knowledge*.json files.
#   * RUST aggregates (S%, full/n, cmp%, medLat, medRsn, A%, grn/n, tok/s, TTFT) are
#     computed OBJECTIVELY from the cargo-oracle records.
#   * Humanitarian per-axis scores come from HUM_SCORES below — set by the Opus 4.8 judge
#     after reading the answers (same judge as the cloud run).
#   * Cost columns are N/A for local models (free inference) -> "local"/"n/a".
#   * reasoning flag (col) is data-derived: median reasoning tokens > 50 -> "yes".
#
# SAFETY: aborts if workbook open (lock file). IDEMPOTENT: skips models already in col A.

import os, sys, json, glob, statistics
import openpyxl
from openpyxl.comments import Comment

DIR  = "/sync/Homie/Obsidian/Primary/claudedocs"
XLSX = os.path.join(DIR, "AI price comparison.xlsx")
LOCK = os.path.join(DIR, ".~lock.AI price comparison.xlsx#")
HERE = os.path.dirname(os.path.abspath(__file__))
RES  = os.path.join(HERE, "results")
GROUP = "local"
PROV  = "local (LM Studio)"

# id -> (short display name, CTX label, quant). Order here drives row order.
META = {
    "ornith-1.0-9b":                          ("Ornith-1.0-9B",      "262k",  "Q4_K_M"),
    "ornith-1.0-35b":                         ("Ornith-1.0-35B",     "262k",  "Q4_K_M"),
    "vibethinker-3b":                         ("VibeThinker-3B",     "131k",  "F16"),
    "essentialai/rnj-1":                      ("RNJ-1",              "33k",   "Q8_0"),
    "mistralai/ministral-3-14b-reasoning":    ("Ministral-3-14B",    "262k",  "Q8_0 (vlm)"),
    "gemma-4-12b-coder-fable5-composer2.5-v1":("Gemma4-12B-coder",   "262k",  "Q8_0"),
    "openai/gpt-oss-20b":                     ("gpt-oss-20b (loc)",  "131k",  "MXFP4"),
    "nvidia/nemotron-3-nano":                 ("Nemotron3-Nano",     "1.05M", "Q3_K_L"),
    "google/gemma-4-26b-a4b-qat":             ("Gemma4-26B-qat",     "262k",  "Q4_0 (vlm)"),
    "qwen/qwen3.6-27b":                       ("Qwen3.6-27B",        "262k",  "Q4_K_M (vlm)"),
    "qwen/qwen3-coder-30b":                   ("Qwen3-Coder-30B",    "262k",  "Q4_K_M"),
    "qwen/qwen3-30b-a3b-2507":                ("Qwen3-30B-A3B",      "262k",  "Q4_K_M"),
    "mistralai/devstral-small-2-2512":        ("Devstral-Small-2",   "393k",  "Q6_K (vlm)"),
    "allenai/olmo-3-32b-think":               ("OLMo-3-32B-think",   "66k",   "Q4_K_M"),
    "qwen/qwq-32b":                           ("QwQ-32B",            "131k",  "Q4_K_M"),
    "bytedance/seed-oss-36b":                 ("Seed-OSS-36B",       "524k",  "Q4_K_M"),
    "zai-org/glm-4.7-flash":                  ("GLM-4.7-Flash",      "203k",  "Q4_K_M"),
    "qwen36-a3b-claude-coder-llama.cpp":      ("Qwen3.6-A3B-coder",  "262k",  "Q4_K_M"),
}

# Humanitarian scores: FILLED BY THE Opus 4.8 JUDGE after reading the answers.
# id -> dict(facts, ideas, fermi, forecast, analysis, empties); None on an axis -> "—" (counts 0).
HUM_SCORES = {
    # batch-2 humanitarian, judged by Opus 4.8 [[29.06.2026]] against the same blind rubric.
    "vibethinker-3b":                          dict(facts=7.0, ideas=6.0, fermi=3.0, forecast=5.0, analysis=7.0, empties=0),
    "essentialai/rnj-1":                       dict(facts=5.5, ideas=5.5, fermi=5.0, forecast=7.0, analysis=7.5, empties=0),
    "mistralai/ministral-3-14b-reasoning":     dict(facts=7.5, ideas=7.0, fermi=5.0, forecast=7.5, analysis=7.5, empties=0),
    "gemma-4-12b-coder-fable5-composer2.5-v1": dict(facts=7.5, ideas=6.5, fermi=9.0, forecast=8.5, analysis=8.0, empties=0),
    "openai/gpt-oss-20b":                      dict(facts=7.5, ideas=6.5, fermi=3.0, forecast=7.0, analysis=8.0, empties=0),
    "nvidia/nemotron-3-nano":                  dict(facts=9.3, ideas=6.5, fermi=5.5, forecast=6.0, analysis=8.0, empties=0),
    "google/gemma-4-26b-a4b-qat":              dict(facts=9.3, ideas=7.5, fermi=9.0, forecast=8.5, analysis=8.0, empties=0),
    "qwen/qwen3.6-27b":                        dict(facts=9.5, ideas=8.0, fermi=9.0, forecast=8.5, analysis=9.0, empties=0),
    "qwen/qwen3-coder-30b":                    dict(facts=9.2, ideas=5.5, fermi=3.5, forecast=7.5, analysis=8.0, empties=0),
    "qwen/qwen3-30b-a3b-2507":                 dict(facts=9.3, ideas=6.5, fermi=5.0, forecast=8.0, analysis=8.0, empties=0),
    "mistralai/devstral-small-2-2512":         dict(facts=7.0, ideas=7.0, fermi=6.0, forecast=7.0, analysis=7.5, empties=0),
    "zai-org/glm-4.7-flash":                   dict(facts=6.5, ideas=6.0, fermi=7.5, forecast=7.5, analysis=7.5, empties=0),
    # qwq-32b: fermi axis exceeded the 20-min deadline locally (~7 tok/s) -> no answer.
    "qwen/qwq-32b":                            dict(facts=9.0, ideas=7.0, fermi=None, forecast=7.5, analysis=8.0, empties=1),
    "ornith-1.0-9b":  dict(facts=9.3, ideas=7.0, fermi=None, forecast=7.5, analysis=9.0, empties=1),
    "ornith-1.0-35b": dict(facts=9.2, ideas=7.5, fermi=9.0, forecast=8.8, analysis=9.4, empties=0),
}


def med(xs):
    xs = [x for x in xs if x is not None]
    return round(statistics.median(xs), 1) if xs else None


def load_all(pattern):
    recs = []
    for f in sorted(glob.glob(os.path.join(RES, pattern))):
        try: recs += json.load(open(f))
        except Exception: pass
    return recs


def rust_aggregate(records, mid):
    single = [r for r in records if r.get("model") == mid and r.get("mode") == "single" and r.get("ok")]
    agentic = [r for r in records if r.get("model") == mid and r.get("mode") == "agentic" and r.get("ok")]
    if not single:
        return None
    # Load-failure guard: a model LM Studio couldn't load returns empty/instant responses
    # (no tokens, no TTFT) — exclude it rather than record a misleading 0%.
    if all((not r.get("tokps")) and r.get("ttft") is None for r in single):
        return {"loadfail": True}
    s_pct = [r["pct"] for r in single]
    out = dict(
        S=round(100 * sum(s_pct) / len(s_pct)),
        full=f"{sum(1 for r in single if r['pct'] >= 0.999)}/{len(single)}",
        cmp=round(100 * sum(1 for r in single if r['compiles']) / len(single)),
        medLat=med([r["latency"] for r in single]),
        medRsn=med([r["reasonTok"] for r in single]),
        tokps=med([r["tokps"] for r in single if r.get("tokps")]),
        ttft=med([r["ttft"] for r in single if r.get("ttft") is not None]),
        rsn="yes" if (med([r["reasonTok"] for r in single]) or 0) > 50 else "no",
    )
    if agentic:
        a_pct = [r["pct"] for r in agentic]
        out["A"] = round(100 * sum(a_pct) / len(a_pct))
        out["grn"] = f"{sum(1 for r in agentic if r.get('visibleGreen'))}/{len(agentic)}"
    else:
        out["A"], out["grn"] = None, None
    return out


def existing_shorts(ws):
    return {(c.value or "").strip() for c in ws["A"] if isinstance(c.value, str)}


def main():
    if os.path.exists(LOCK):
        sys.exit("ABORT: workbook open in LibreOffice (lock present). Close it and re-run.")
    rust_recs = load_all("*rust*.json")
    know_recs = load_all("*knowledge*.json")

    wb = openpyxl.load_workbook(XLSX)
    rust, hum = wb["RUST"], wb["humanitarian"]
    have_r, have_h = existing_shorts(rust), existing_shorts(hum)

    for mid, (short, ctx, quant) in META.items():
        agg = rust_aggregate(rust_recs, mid)
        if agg and agg.get("loadfail"):
            print(f"RUST  LOAD-FAILED (skip): {short} — LM Studio could not load this model")
            agg = None
        desc = f"{short} — локально в LM Studio на gaming-pc, {quant}, {ctx} ctx. Прогон [[28.06.2026]]."
        # ---- RUST sheet (20 cols A..T) ----
        if agg and short not in have_r:
            rn = rust.max_row + 1
            J_ = agg["A"] if agg["A"] is not None else "—"
            K_ = agg["grn"] if agg["grn"] is not None else "—"
            row = [short, GROUP, "local", agg["S"], agg["full"], agg["cmp"], "local",
                   agg["medLat"], agg["medRsn"], J_, K_, "local", "n/a", "n/a",
                   ctx, "yes", PROV, agg["rsn"], agg["tokps"], agg["ttft"]]
            rust.append(row)
            c = Comment(desc, "bench"); c.width, c.height = 340, 90
            rust.cell(row=rn, column=1).comment = c
            print(f"RUST  + {short}: S%={agg['S']} full={agg['full']} cmp%={agg['cmp']} A%={agg['A']} grn={agg['grn']} medLat={agg['medLat']} medRsn={agg['medRsn']} tok/s={agg['tokps']} ttft={agg['ttft']} rsn={agg['rsn']}")
        elif short in have_r:
            print(f"RUST  skip (present): {short}")
        elif not agg:
            print(f"RUST  no data yet: {short}")

        # ---- humanitarian sheet (18 cols A..R) ----
        sc = HUM_SCORES.get(mid)
        if sc and short not in have_h:
            rn = hum.max_row + 1
            axes = [sc.get(k) for k in ("facts", "ideas", "fermi", "forecast", "analysis")]
            vals = [(0 if a is None else a) for a in axes]
            avg = round(sum(vals) / len(vals), 2)
            disp = [("—" if a is None else a) for a in axes]
            kc = [r for r in know_recs if r.get("model") == mid and r.get("ok")]
            ktok = med([r["tokps"] for r in kc if r.get("tokps")]) or (agg["tokps"] if agg else None)
            kttft = med([r["ttft"] for r in kc if r.get("ttft") is not None]) or (agg["ttft"] if agg else None)
            krsn = "yes" if (med([r.get("reasonTok", 0) for r in kc]) or 0) > 50 else "no"
            row = [short, GROUP] + disp + [avg, "local", sc.get("empties", 0), "n/a", "n/a",
                   ctx, "yes", PROV, krsn, ktok, kttft]
            hum.append(row)
            c = Comment(desc, "bench"); c.width, c.height = 340, 90
            hum.cell(row=rn, column=1).comment = c
            print(f"HUM   + {short}: facts={disp[0]} ideas={disp[1]} fermi={disp[2]} forecast={disp[3]} analysis={disp[4]} avg={avg} empties={sc.get('empties',0)} tok/s={ktok}")
        elif short in have_h:
            print(f"HUM   skip (present): {short}")
        elif not sc:
            print(f"HUM   pending judge: {short}")

    wb.save(XLSX)
    print("Saved.")


if __name__ == "__main__":
    main()
